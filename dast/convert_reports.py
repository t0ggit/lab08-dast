import json
import os
import sys
from datetime import datetime

print("[debug] python:", sys.executable)

try:
    from odf.opendocument import OpenDocumentText
    from odf.text import P, H
    from odf.style import Style, TextProperties

    HAS_ODF = True
    print("[debug] odf imported OK")
except ImportError as e:
    HAS_ODF = False
    print("[debug] odf import error:", e)
    print("[!] Skipping ODT generation: odfpy not installed")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[!] WARNING: openpyxl не установлен. XLSX не будет генерироваться.")


def parse_zap_json(json_path: str) -> dict:
    with open(json_path, "r") as f:
        data = json.load(f)

    alerts = data.get("site", [{}])[0].get("alerts", [])

    return {
        "scan_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_alerts": len(alerts),
        "alerts": alerts,
    }


def generate_odt(zap_dict, output_path: str) -> None:
    if not HAS_ODF:
        print("[!] Skipping ODT generation: odfpy not installed")
        return

    doc = OpenDocumentText()

    title = H(outlinelevel=1)
    title.addText("OWASP ZAP DAST Report")
    doc.text.addElement(title)

    meta_p = P()
    meta_p.addText(f"Scan Date: {zap_dict['scan_date']}")
    doc.text.addElement(meta_p)

    summary_p = P()
    summary_p.addText(f"Total Alerts: {zap_dict['total_alerts']}")
    doc.text.addElement(summary_p)

    if zap_dict["alerts"]:
        table_header = H(outlinelevel=2)
        table_header.addText("Vulnerabilities")
        doc.text.addElement(table_header)

        for i, alert in enumerate(zap_dict["alerts"][:20], 1):  # первые 20
            alert_p = P()
            alert_p.addText(
                f"{i}. {alert.get('name', 'Unknown')} "
                f"[{alert.get('risk', 'N/A')}] "
                f"@ {alert.get('url', 'N/A')}"
            )
            doc.text.addElement(alert_p)

    doc.save(output_path)
    print(f"[+] ODT report saved: {output_path}")


def generate_xlsx(zap_dict, output_path: str) -> None:
    if not HAS_OPENPYXL:
        print("[!] Skipping XLSX generation: openpyxl not installed")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "ZAP Scan Results"

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["#", "Alert Name", "Risk Level", "URL", "Description", "Solution"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for row_idx, alert in enumerate(zap_dict["alerts"][:100], 2):
        ws.cell(row=row_idx, column=1).value = row_idx - 1
        ws.cell(row=row_idx, column=2).value = alert.get("name", "Unknown")
        ws.cell(row=row_idx, column=3).value = alert.get("risk", "Informational")
        ws.cell(row=row_idx, column=4).value = alert.get("url", "")
        ws.cell(row=row_idx, column=5).value = alert.get("description", "")
        ws.cell(row=row_idx, column=6).value = alert.get("solution", "")

        ws.row_dimensions[row_idx].height = None
        for col in [2, 5, 6]:
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 40

    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"[+] XLSX report saved: {output_path}")


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python3 convert_reports.py <timestamp>")
        sys.exit(1)

    timestamp = sys.argv[1]
    json_path = f"zap-report-{timestamp}.json"

    if not os.path.exists(json_path):
        print(f"[!] JSON report not found: {json_path}")
        sys.exit(1)

    print(f"[*] Parsing ZAP JSON report: {json_path}")
    zap_dict = parse_zap_json(json_path)
    print(f"[i] Found {zap_dict['total_alerts']} alerts")

    odt_dir = "odt"
    xlsx_dir = "xlsx"
    os.makedirs(odt_dir, exist_ok=True)
    os.makedirs(xlsx_dir, exist_ok=True)

    odt_path = os.path.join(odt_dir, f"zap-report-{timestamp}.odt")
    xlsx_path = os.path.join(xlsx_dir, f"zap-report-{timestamp}.xlsx")

    generate_odt(zap_dict, odt_path)
    generate_xlsx(zap_dict, xlsx_path)

    print("[+] Report conversion completed!")


if __name__ == "__main__":
    main()
